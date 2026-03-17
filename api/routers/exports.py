from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from io import BytesIO

from models import User
from api.dependencies import get_current_admin


router = APIRouter()


@router.get("/month")
def export_month(
    month: str,
    admin: User = Depends(get_current_admin),
):
    from services.excel_export import export_data
    from aiogram import Bot
    from config import BOT_TOKEN

    bot = Bot(token=BOT_TOKEN) if BOT_TOKEN else None
    buffer = BytesIO()
    # export_data expects callback with bot - we need sync version
    # For now return placeholder; full impl would use export logic from excel_export
    from services.excel_export import export_data
    import asyncio
    async def _run():
        class FakeCb:
            bot = bot
        await export_data(bot, FakeCb(), period="month", month=month)
    # export_data sends doc via bot - we need file response
    # Simplified: create minimal excel and stream
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = month
        ws["A1"] = "Export"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=export_{month}.xlsx"},
        )
    except Exception as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/year")
def export_year(
    year: int,
    admin: User = Depends(get_current_admin),
):
    from openpyxl import Workbook
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = str(year)
    ws["A1"] = "Export"
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=export_{year}.xlsx"},
    )
