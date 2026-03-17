import os
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.drawing.image import Image
from datetime import datetime
from aiogram import types
from config import MAX_MESSAGE_LENGTH
from models import RegistrationStatus
from database import SessionLocal
from sqlalchemy.exc import SQLAlchemyError

logger = logging.getLogger(__name__)

def split_text(text: str, max_length: int = MAX_MESSAGE_LENGTH) -> list:
    messages = []
    while len(text) > max_length:
        split_index = text.rfind('\n', 0, max_length)
        if split_index == -1:
            split_index = max_length
        messages.append(text[:split_index])
        text = text[split_index:]
    messages.append(text)
    return messages

async def export_data(bot, callback_query: types.CallbackQuery, period, month=None, year=None):
    from models import Registration, Tournament, User
    session = SessionLocal()
    try:
        # построить запрос
        query = session.query(Registration).join(Tournament).join(User)
        if period == 'month' and month:
            query = query.filter(Tournament.month == month)
        elif period == 'year' and year:
            start = datetime(year, 1, 1).date()
            end   = datetime(year,12,31).date()
            query = query.filter(Tournament.date.between(start, end))
        elif period == 'season':
            today = datetime.now().date()
            if today.month >= 9:
                sy, ey = today.year, today.year + 1
            else:
                sy, ey = today.year - 1, today.year
            start_date = datetime(sy,9,1).date()
            end_date = datetime(ey,5,31).date()
            logger.info(f"Экспорт за сезон {sy}-{ey}: {start_date} - {end_date}")
            query = query.filter(
                Tournament.date.between(start_date, end_date)
            )
        # else 'all' — без фильтра

        regs = query.all()
        if not regs:
            await bot.send_message(callback_query.from_user.id, "❌ Нет данных для выбранного периода.")
            return

        # сортировка
        regs.sort(key=lambda r: r.tournament.date)

        # Создаем красивую книгу
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Создаем листы
        ws_main = wb.create_sheet("📊 Заявки", 0)
        ws_stats = wb.create_sheet("📈 Статистика", 1)
        ws_summary = wb.create_sheet("📋 Сводка", 2)

        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === ЛИСТ 1: ЗАЯВКИ ===
        # Заголовок с логотипом
        ws_main.merge_cells('A1:F1')
        title_cell = ws_main['A1']
        title_cell.value = f"🏆 ОТЧЕТ ПО ЗАЯВКАМ НА ТУРНИРЫ - {datetime.now().strftime('%d.%m.%Y')}"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill('solid', fgColor='1F4E79')
        
        # Подзаголовок с информацией о периоде
        ws_main.merge_cells('A2:F2')
        subtitle_cell = ws_main['A2']
        period_text = f"📅 Период: {period.upper()} {month or year or 'ВСЕ ВРЕМЕНА'}"
        subtitle_cell.value = period_text
        subtitle_cell.font = Font(name='Arial', size=12, italic=True, color='2F5597')
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        subtitle_cell.fill = PatternFill('solid', fgColor='F2F2F2')

        # Заголовки таблицы с улучшенным дизайном
        headers = ['📅 Дата турнира', '🏆 Название турнира', '👤 Судья', '⚙️ Функция', '🏅 Категория', '📊 Статус']
        for idx, header in enumerate(headers, 1):
            cell = ws_main.cell(row=4, column=idx, value=header)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )

        # Данные
        row = 5
        for r in regs:
            t, u = r.tournament, r.user
            
            # Форматирование статуса
            status_emoji = {
                RegistrationStatus.APPROVED: '✅ Утверждено',
                RegistrationStatus.REJECTED: '❌ Отклонено',
                RegistrationStatus.PENDING: '⏳ Ожидает'
            }.get(r.status, r.status)
            
            # Цветовая схема для статусов
            status_fill = {
                RegistrationStatus.APPROVED: PatternFill('solid', fgColor='D5E8D4'),
                RegistrationStatus.REJECTED: PatternFill('solid', fgColor='F8CECC'),
                RegistrationStatus.PENDING: PatternFill('solid', fgColor='FFF2CC')
            }.get(r.status)

            data_row = [
                t.date.strftime('%d.%m.%Y'),
                t.name,
                f"{u.first_name} {u.last_name}",
                u.function,
                u.category,
                status_emoji
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws_main.cell(row=row, column=col, value=value)
                cell.font = Font(name='Arial', size=10, bold=(col == 3))  # Жирный шрифт для имен судей
                cell.alignment = data_alignment
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if status_fill:
                    cell.fill = status_fill
                
                # Специальное форматирование для даты
                if col == 1:
                    cell.number_format = 'DD.MM.YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1

        # Настройка размеров колонок
        column_widths = [15, 30, 25, 20, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws_main.column_dimensions[get_column_letter(i)].width = width

        # Заморозка заголовков
        ws_main.freeze_panes = 'A5'
        
        # Добавляем итоговую статистику в конец листа
        summary_row = row + 2
        ws_main.merge_cells(f'A{summary_row}:F{summary_row}')
        summary_cell = ws_main[f'A{summary_row}']
        summary_cell.value = "📊 ИТОГОВАЯ СТАТИСТИКА"
        summary_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        summary_cell.fill = PatternFill('solid', fgColor='1F4E79')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Статистика по статусам
        stats_row = summary_row + 1
        status_stats = {}
        for r in regs:
            status = r.status
            status_stats[status] = status_stats.get(status, 0) + 1
        
        ws_main[f'A{stats_row}'].value = "✅ Утверждено:"
        ws_main[f'B{stats_row}'].value = status_stats.get(RegistrationStatus.APPROVED, 0)
        ws_main[f'C{stats_row}'].value = "❌ Отклонено:"
        ws_main[f'D{stats_row}'].value = status_stats.get(RegistrationStatus.REJECTED, 0)
        ws_main[f'E{stats_row}'].value = "⏳ Ожидает:"
        ws_main[f'F{stats_row}'].value = status_stats.get(RegistrationStatus.PENDING, 0)
        
        # Форматирование статистики
        for col in range(1, 7):
            cell = ws_main.cell(row=stats_row, column=col)
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.fill = PatternFill('solid', fgColor='F2F2F2')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # === ЛИСТ 2: СТАТИСТИКА ===
        # Заголовок с градиентом
        ws_stats.merge_cells('A1:H1')
        title_cell = ws_stats['A1']
        title_cell.value = "📈 СТАТИСТИКА ПО СУДЬЯМ"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill('solid', fgColor='1F4E79')
        
        # Подзаголовок с датой генерации
        ws_stats.merge_cells('A2:H2')
        subtitle_cell = ws_stats['A2']
        subtitle_cell.value = f"📊 Сгенерировано: {datetime.now().strftime('%d.%m.%Y в %H:%M')}"
        subtitle_cell.font = Font(name='Arial', size=10, italic=True, color='666666')
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        subtitle_cell.fill = PatternFill('solid', fgColor='F8F9FA')

        # Сбор статистики с реальными данными об оплате
        stats = {}
        for r in regs:
            name = f"{r.user.first_name} {r.user.last_name}"
            if name not in stats:
                stats[name] = {'Всего': 0, 'Утверждено': 0, 'Отклонено': 0, 'Ожидает': 0, 'Заработок': 0, 'Оплаченных_турниров': 0}
            
            stats[name]['Всего'] += 1
            if r.status == RegistrationStatus.APPROVED:
                stats[name]['Утверждено'] += 1
            elif r.status == RegistrationStatus.REJECTED:
                stats[name]['Отклонено'] += 1
            else:
                stats[name]['Ожидает'] += 1
        
        # Получаем реальные данные об оплате
        from models import JudgePayment
        from sqlalchemy import func, and_
        
        # Получаем заработок каждого судьи за выбранный период
        earnings_query = session.query(
            User.first_name,
            User.last_name,
            func.sum(JudgePayment.amount).label('total_amount'),
            func.count(JudgePayment.payment_id).label('paid_tournaments')
        ).join(
            JudgePayment, User.user_id == JudgePayment.user_id
        ).join(
            Tournament, JudgePayment.tournament_id == Tournament.tournament_id
        ).filter(
            and_(
                JudgePayment.is_paid == True,
                JudgePayment.amount.isnot(None)
            )
        )
        
        # Применяем фильтры по периоду
        if period == 'month' and month:
            earnings_query = earnings_query.filter(Tournament.month == month)
        elif period == 'year' and year:
            start = datetime(year, 1, 1).date()
            end = datetime(year, 12, 31).date()
            earnings_query = earnings_query.filter(Tournament.date.between(start, end))
        elif period == 'season':
            today = datetime.now().date()
            if today.month >= 9:
                sy, ey = today.year, today.year + 1
            else:
                sy, ey = today.year - 1, today.year
            start_date = datetime(sy, 9, 1).date()
            end_date = datetime(ey, 5, 31).date()
            earnings_query = earnings_query.filter(Tournament.date.between(start_date, end_date))
        
        earnings_data = earnings_query.group_by(User.user_id).all()
        
        # Обновляем статистику реальными данными об оплате
        for first_name, last_name, amount, paid_tournaments in earnings_data:
            name = f"{first_name} {last_name}"
            if name in stats:
                stats[name]['Заработок'] = amount or 0
                stats[name]['Оплаченных_турниров'] = paid_tournaments or 0

        sorted_stats = sorted(stats.items(), key=lambda x: x[1]['Всего'], reverse=True)

        # Заголовки статистики с улучшенным дизайном
        stat_headers = ['👤 Судья', '📊 Всего', '✅ Утверждено', '❌ Отклонено', '⏳ Ожидает', '📈 % Успеха', '💰 Заработок', '🏆 Рейтинг']
        for i, header in enumerate(stat_headers, 1):
            cell = ws_stats.cell(row=4, column=i, value=header)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )

        # Данные статистики с рейтинговой системой
        row = 5
        for rank, (name, st) in enumerate(sorted_stats, 1):
            success_rate = (st['Утверждено'] / st['Всего']) * 100 if st['Всего'] > 0 else 0
            earnings = st['Заработок']  # Реальные данные об оплате
            
            # Рейтинговая система
            rating = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉" if rank == 3 else f"#{rank}"
            
            data_row = [
                name,
                st['Всего'],
                st['Утверждено'],
                st['Отклонено'],
                st['Ожидает'],
                f"{success_rate:.1f}%",
                f"{earnings:,.2f} ₽" if earnings > 0 else "0 ₽",
                rating
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws_stats.cell(row=row, column=col, value=value)
                
                # Специальное форматирование для разных колонок
                if col == 1:  # Имя судьи
                    cell.font = Font(name='Arial', size=10, bold=True)
                    cell.fill = PatternFill('solid', fgColor='E7F3FF')
                elif col == 6:  # Процент успеха
                    cell.font = Font(name='Arial', size=10, bold=True)
                    if success_rate >= 80:
                        cell.fill = PatternFill('solid', fgColor='D5E8D4')  # Зеленый
                    elif success_rate >= 60:
                        cell.fill = PatternFill('solid', fgColor='FFF2CC')  # Желтый
                    else:
                        cell.fill = PatternFill('solid', fgColor='F8CECC')  # Красный
                elif col == 7:  # Заработок
                    cell.font = Font(name='Arial', size=10, bold=True)
                    cell.fill = PatternFill('solid', fgColor='F0F8FF')
                elif col == 8:  # Рейтинг
                    cell.font = Font(name='Arial', size=12, bold=True)
                    cell.fill = PatternFill('solid', fgColor='FFFACD')
                else:
                    cell.font = Font(name='Arial', size=10)
                    cell.fill = PatternFill('solid', fgColor='FFFFFF')
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1

        # Настройка размеров колонок для статистики
        stat_widths = [25, 8, 10, 10, 8, 12, 15, 15]
        for i, width in enumerate(stat_widths, 1):
            ws_stats.column_dimensions[get_column_letter(i)].width = width

        ws_stats.freeze_panes = 'A5'
        
        # Добавляем итоговую статистику в конец листа статистики
        summary_row = row + 2
        ws_stats.merge_cells(f'A{summary_row}:H{summary_row}')
        summary_cell = ws_stats[f'A{summary_row}']
        summary_cell.value = "🏆 ТОП-3 СУДЕЙ ПО АКТИВНОСТИ"
        summary_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        summary_cell.fill = PatternFill('solid', fgColor='1F4E79')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Показываем топ-3
        top3_row = summary_row + 1
        for i, (name, st) in enumerate(sorted_stats[:3], 1):
            medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉"
            ws_stats[f'A{top3_row}'].value = f"{medal} {i}. {name}"
            ws_stats[f'B{top3_row}'].value = f"{st['Всего']} заявок"
            ws_stats[f'C{top3_row}'].value = f"{(st['Утверждено'] / st['Всего'] * 100):.1f}% успеха"
            ws_stats[f'D{top3_row}'].value = f"{st['Заработок']:,.2f} ₽" if st['Заработок'] > 0 else "0 ₽"
            
            # Форматирование топ-3
            for col in range(1, 5):
                cell = ws_stats.cell(row=top3_row, column=col)
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.fill = PatternFill('solid', fgColor='F0F8FF')
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            top3_row += 1

        # === ЛИСТ 3: СВОДКА ===
        # Заголовок с логотипом
        ws_summary.merge_cells('A1:F1')
        title_cell = ws_summary['A1']
        title_cell.value = "📋 СВОДНАЯ ИНФОРМАЦИЯ И АНАЛИТИКА"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill('solid', fgColor='1F4E79')
        
        # Подзаголовок
        ws_summary.merge_cells('A2:F2')
        subtitle_cell = ws_summary['A2']
        subtitle_cell.value = f"📊 Детальный анализ за период: {period.upper()} {month or year or 'ВСЕ ВРЕМЕНА'}"
        subtitle_cell.font = Font(name='Arial', size=12, italic=True, color='2F5597')
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        subtitle_cell.fill = PatternFill('solid', fgColor='F2F2F2')

        # Общая статистика
        total_registrations = len(regs)
        total_approved = sum(1 for r in regs if r.status == RegistrationStatus.APPROVED)
        total_rejected = sum(1 for r in regs if r.status == RegistrationStatus.REJECTED)
        total_pending = sum(1 for r in regs if r.status == RegistrationStatus.PENDING)
        
        unique_judges = len(set(f"{r.user.first_name} {r.user.last_name}" for r in regs))
        unique_tournaments = len(set(r.tournament.tournament_id for r in regs))
        
        approval_rate = (total_approved / total_registrations * 100) if total_registrations > 0 else 0
        
        # Реальный общий заработок
        total_earnings = sum(st['Заработок'] for st in stats.values())
        total_paid_tournaments = sum(st['Оплаченных_турниров'] for st in stats.values())

        # Создаем красивую таблицу сводки
        summary_data = [
            ['📊 Общее количество заявок:', total_registrations, 'шт.'],
            ['✅ Утверждено:', total_approved, 'шт.'],
            ['❌ Отклонено:', total_rejected, 'шт.'],
            ['⏳ Ожидает рассмотрения:', total_pending, 'шт.'],
            ['👥 Уникальных судей:', unique_judges, 'чел.'],
            ['🏆 Уникальных турниров:', unique_tournaments, 'шт.'],
            ['📈 Процент утверждения:', f"{approval_rate:.1f}%", ''],
            ['💰 Общий заработок:', f"{total_earnings:,.2f} ₽", ''],
            ['💳 Оплаченных турниров:', total_paid_tournaments, 'шт.'],
            ['📅 Период отчета:', f"{period} {month or year or ''}", ''],
            ['🕐 Время генерации:', datetime.now().strftime('%d.%m.%Y %H:%M'), '']
        ]

        # Заголовки таблицы сводки
        ws_summary.cell(row=4, column=1, value="📋 Показатель").font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        ws_summary.cell(row=4, column=2, value="📊 Значение").font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        ws_summary.cell(row=4, column=3, value="📏 Единица").font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        
        for col in range(1, 4):
            cell = ws_summary.cell(row=4, column=col)
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )

        for i, (label, value, unit) in enumerate(summary_data, 5):
            # Показатель
            ws_summary.cell(row=i, column=1, value=label).font = Font(name='Arial', size=11, bold=True)
            ws_summary.cell(row=i, column=1).fill = PatternFill('solid', fgColor='E7F3FF')
            ws_summary.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
            
            # Значение
            ws_summary.cell(row=i, column=2, value=value).font = Font(name='Arial', size=11, bold=True)
            ws_summary.cell(row=i, column=2).fill = PatternFill('solid', fgColor='FFFFFF')
            ws_summary.cell(row=i, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            # Единица измерения
            ws_summary.cell(row=i, column=3, value=unit).font = Font(name='Arial', size=10)
            ws_summary.cell(row=i, column=3).fill = PatternFill('solid', fgColor='F8F9FA')
            ws_summary.cell(row=i, column=3).alignment = Alignment(horizontal='center', vertical='center')
            
            # Границы
            for col in range(1, 4):
                cell = ws_summary.cell(row=i, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Настройка размеров
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 10

        # === ГРАФИКИ ===
        # Создаем данные для диаграмм в ячейках
        chart_start_row = len(summary_data) + 8  # После сводной таблицы
        
        # Заголовок для данных диаграммы
        ws_summary.cell(row=chart_start_row, column=1, value="Статус").font = Font(bold=True)
        ws_summary.cell(row=chart_start_row, column=2, value="Количество").font = Font(bold=True)
        
        # Данные для круговой диаграммы
        chart_data = [
            ('✅ Утверждено', total_approved),
            ('❌ Отклонено', total_rejected),
            ('⏳ Ожидает', total_pending)
        ]
        
        # Записываем данные в ячейки
        for i, (status, count) in enumerate(chart_data, 1):
            ws_summary.cell(row=chart_start_row + i, column=1, value=status)
            ws_summary.cell(row=chart_start_row + i, column=2, value=count)
        
        # Создаем круговую диаграмму
        if any(count > 0 for _, count in chart_data):
            pie_chart = PieChart()
            pie_chart.title = "📊 Распределение заявок по статусам"
            
            # Создаем Reference для данных
            from openpyxl.chart import Reference
            data_ref = Reference(ws_summary, min_col=2, min_row=chart_start_row + 1, max_row=chart_start_row + len(chart_data))
            cats_ref = Reference(ws_summary, min_col=1, min_row=chart_start_row + 1, max_row=chart_start_row + len(chart_data))
            
            pie_chart.add_data(data_ref)
            pie_chart.set_categories(cats_ref)
            pie_chart.style = 10
            
            ws_summary.add_chart(pie_chart, "E5")

        # Улучшенная столбчатая диаграмма активности судей
        if len(sorted_stats) > 0:
            # Создаем данные для столбчатой диаграммы в ячейках
            bar_chart_start_row = len(sorted_stats) + 8  # После таблицы статистики
            
            # Заголовок для данных диаграммы
            ws_stats.cell(row=bar_chart_start_row, column=1, value="Судья").font = Font(bold=True)
            ws_stats.cell(row=bar_chart_start_row, column=2, value="Заявки").font = Font(bold=True)
            
            # Берем топ-10 судей
            top_judges = sorted_stats[:10]
            judges_names = [name[:12] + '...' if len(name) > 12 else name for name, _ in top_judges]
            judges_counts = [data['Всего'] for _, data in top_judges]
            
            # Записываем данные в ячейки
            for i, (name, count) in enumerate(zip(judges_names, judges_counts), 1):
                ws_stats.cell(row=bar_chart_start_row + i, column=1, value=name)
                ws_stats.cell(row=bar_chart_start_row + i, column=2, value=count)
            
            # Создаем столбчатую диаграмму
            if judges_counts:
                bar_chart = BarChart()
                bar_chart.title = "🏆 ТОП-10 СУДЕЙ ПО АКТИВНОСТИ"
                bar_chart.y_axis.title = 'Количество заявок'
                bar_chart.x_axis.title = 'Судьи'
                
                # Создаем Reference для данных
                from openpyxl.chart import Reference
                data_ref = Reference(ws_stats, min_col=2, min_row=bar_chart_start_row + 1, max_row=bar_chart_start_row + len(judges_counts))
                cats_ref = Reference(ws_stats, min_col=1, min_row=bar_chart_start_row + 1, max_row=bar_chart_start_row + len(judges_counts))
                
                bar_chart.add_data(data_ref)
                bar_chart.set_categories(cats_ref)
                bar_chart.style = 10
                
                ws_stats.add_chart(bar_chart, "J5")

        # Сохранить и отправить
        tmp = f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(tmp)
        
        with open(tmp, 'rb') as f:
            await bot.send_document(
                callback_query.from_user.id, 
                f,
                caption=f"🎉 <b>ПРОФЕССИОНАЛЬНЫЙ ОТЧЕТ ГОТОВ!</b> 🎉\n\n"
                       f"📅 <b>Период:</b> {period.upper()} {month or year or 'ВСЕ ВРЕМЕНА'}\n"
                       f"📈 <b>Всего заявок:</b> {total_registrations:,}\n"
                       f"✅ <b>Утверждено:</b> {total_approved:,}\n"
                       f"❌ <b>Отклонено:</b> {total_rejected:,}\n"
                       f"⏳ <b>Ожидает:</b> {total_pending:,}\n"
                       f"📊 <b>Процент успеха:</b> {approval_rate:.1f}%\n"
                       f"👥 <b>Уникальных судей:</b> {unique_judges}\n"
                       f"🏆 <b>Уникальных турниров:</b> {unique_tournaments}\n"
                       f"💰 <b>Общий заработок:</b> {total_earnings:,.2f} ₽\n"
                       f"💳 <b>Оплаченных турниров:</b> {total_paid_tournaments}\n\n"
                       f"📋 <b>Отчет содержит:</b>\n"
                       f"• Детальную таблицу заявок\n"
                       f"• Статистику по судьям с рейтингом\n"
                       f"• Сводную аналитику\n"
                       f"• Интерактивные графики\n"
                       f"• Топ-3 самых активных судей",
                parse_mode="HTML"
            )
        
        os.remove(tmp)
        await bot.send_message(
            callback_query.from_user.id, 
            "✨ <b>Отчет создан с использованием профессионального дизайна!</b>\n\n"
            "🎨 <b>Особенности отчета:</b>\n"
            "• Красивое форматирование с цветами\n"
            "• Рейтинговая система судей\n"
            "• Интерактивные диаграммы\n"
            "• Детальная аналитика\n"
            "• Профессиональный стиль",
            parse_mode="HTML"
        )

    except SQLAlchemyError as e:
        await bot.send_message(callback_query.from_user.id, f"❌ Ошибка при экспорте:\n{e}")
    finally:
        session.close()
